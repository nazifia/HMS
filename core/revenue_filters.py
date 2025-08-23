"""
Advanced Filtering and Export Utilities for Revenue Point Analysis
Provides comprehensive filtering options and export capabilities
while maintaining integration with existing HMS systems.
"""

from django import forms
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import json
from io import StringIO, BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .revenue_point_analyzer import RevenuePointFilterHelper
from billing.models import Payment as BillingPayment
from pharmacy_billing.models import Payment as PharmacyPayment
from patients.models import WalletTransaction


class AdvancedRevenueFilterForm(forms.Form):
    """
    Advanced filtering form for revenue point analysis
    """
    
    # Date filtering options
    date_filter = forms.ChoiceField(
        choices=RevenuePointFilterHelper.get_filter_options(),
        initial='current_month',
        widget=forms.Select(attrs={
            'class': 'form-control',
            'id': 'date_filter'
        })
    )
    
    start_date = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={
            'type': 'date',
            'class': 'form-control',
            'id': 'start_date'
        })
    )
    
    end_date = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={
            'type': 'date',
            'class': 'form-control',
            'id': 'end_date'
        })
    )
    
    # Department filtering
    department_filter = forms.ChoiceField(
        choices=RevenuePointFilterHelper.get_department_filter_options(),
        initial='all',
        widget=forms.Select(attrs={
            'class': 'form-control',
            'id': 'department_filter'
        })
    )
    
    # Payment method filtering
    payment_method_filter = forms.ChoiceField(
        choices=RevenuePointFilterHelper.get_payment_method_options(),
        initial='all',
        widget=forms.Select(attrs={
            'class': 'form-control',
            'id': 'payment_method_filter'
        })
    )
    
    # Revenue range filtering
    min_revenue = forms.DecimalField(
        required=False,
        min_value=0,
        widget=forms.NumberInput(attrs={
            'class': 'form-control',
            'placeholder': 'Minimum revenue',
            'step': '0.01'
        })
    )
    
    max_revenue = forms.DecimalField(
        required=False,
        min_value=0,
        widget=forms.NumberInput(attrs={
            'class': 'form-control',
            'placeholder': 'Maximum revenue',
            'step': '0.01'
        })
    )
    
    # Include/exclude options
    include_refunds = forms.BooleanField(
        required=False,
        initial=True,
        widget=forms.CheckboxInput(attrs={
            'class': 'form-check-input'
        })
    )
    
    include_cancelled = forms.BooleanField(
        required=False,
        initial=False,
        widget=forms.CheckboxInput(attrs={
            'class': 'form-check-input'
        })
    )
    
    # Grouping options
    GROUP_BY_CHOICES = [
        ('none', 'No Grouping'),
        ('department', 'By Department'),
        ('payment_method', 'By Payment Method'),
        ('service_category', 'By Service Category'),
        ('day', 'By Day'),
        ('week', 'By Week'),
        ('month', 'By Month')
    ]
    
    group_by = forms.ChoiceField(
        choices=GROUP_BY_CHOICES,
        initial='none',
        widget=forms.Select(attrs={
            'class': 'form-control'
        })
    )
    
    # Sorting options
    SORT_BY_CHOICES = [
        ('revenue_desc', 'Revenue (High to Low)'),
        ('revenue_asc', 'Revenue (Low to High)'),
        ('transactions_desc', 'Transactions (High to Low)'),
        ('transactions_asc', 'Transactions (Low to High)'),
        ('date_desc', 'Date (Newest First)'),
        ('date_asc', 'Date (Oldest First)'),
        ('department', 'Department Name')
    ]
    
    sort_by = forms.ChoiceField(
        choices=SORT_BY_CHOICES,
        initial='revenue_desc',
        widget=forms.Select(attrs={
            'class': 'form-control'
        })
    )
    
    def clean(self):
        cleaned_data = super().clean()
        date_filter = cleaned_data.get('date_filter')
        start_date = cleaned_data.get('start_date')
        end_date = cleaned_data.get('end_date')
        
        # Validate custom date range
        if date_filter == 'custom_range':
            if not start_date or not end_date:
                raise forms.ValidationError(
                    "Start date and end date are required for custom range."
                )
            if start_date > end_date:
                raise forms.ValidationError(
                    "Start date must be before end date."
                )
            if (end_date - start_date).days > 365:
                raise forms.ValidationError(
                    "Date range cannot exceed 365 days."
                )
        
        # Validate revenue range
        min_revenue = cleaned_data.get('min_revenue')
        max_revenue = cleaned_data.get('max_revenue')
        
        if min_revenue and max_revenue and min_revenue > max_revenue:
            raise forms.ValidationError(
                "Minimum revenue must be less than maximum revenue."
            )
        
        return cleaned_data


class RevenueExportUtility:
    """
    Utility class for exporting revenue data in various formats
    """
    
    def __init__(self, breakdown_data, date_range, filters=None):
        self.breakdown_data = breakdown_data
        self.date_range = date_range
        self.filters = filters or {}
    
    def export_csv(self):
        """
        Export revenue breakdown as CSV
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Header information
        writer.writerow(['HMS Revenue Point Breakdown Analysis'])
        writer.writerow(['Date Range', f"{self.date_range['start_date']} to {self.date_range['end_date']}"])
        writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(['Filters Applied', self._format_filters()])
        writer.writerow([])
        
        # Summary section
        writer.writerow(['REVENUE SUMMARY'])
        writer.writerow(['Category', 'Revenue (₦)', 'Percentage (%)'])
        
        summary = self.breakdown_data.get('summary_by_category', {})
        for category, data in summary.items():
            if category != 'grand_total':
                writer.writerow([
                    category.replace('_', ' ').title(),
                    f"{data.get('revenue', 0):.2f}",
                    f"{data.get('percentage', 0):.2f}"
                ])
        
        writer.writerow(['Total Revenue', f"{summary.get('grand_total', 0):.2f}", '100.00'])
        writer.writerow([])
        
        # Clinical Services Breakdown
        writer.writerow(['CLINICAL SERVICES BREAKDOWN'])
        writer.writerow(['Service', 'Revenue (₦)', 'Transactions', 'Average Transaction (₦)'])
        
        clinical = self.breakdown_data.get('clinical_services', {})
        for service, data in clinical.items():
            writer.writerow([
                service.title(),
                f"{data.get('revenue', 0):.2f}",
                data.get('transactions', 0),
                f"{data.get('avg_transaction', 0):.2f}"
            ])
        
        writer.writerow([])
        
        # Support Services Breakdown
        writer.writerow(['SUPPORT SERVICES BREAKDOWN'])
        writer.writerow(['Service', 'Revenue (₦)', 'Transactions', 'Average Transaction (₦)'])
        
        support = self.breakdown_data.get('support_services', {})
        for service, data in support.items():
            writer.writerow([
                service.title(),
                f"{data.get('revenue', 0):.2f}",
                data.get('transactions', 0),
                f"{data.get('avg_transaction', 0):.2f}"
            ])
        
        writer.writerow([])
        
        # Specialty Departments
        writer.writerow(['SPECIALTY DEPARTMENTS BREAKDOWN'])
        writer.writerow(['Department', 'Revenue (₦)', 'Records', 'Revenue per Record (₦)'])
        
        specialty = self.breakdown_data.get('specialty_departments', {})
        for dept, data in specialty.items():
            avg_per_record = 0
            if data.get('records', 0) > 0:
                avg_per_record = data.get('revenue', 0) / data.get('records', 1)
            
            writer.writerow([
                dept.replace('_', ' ').title(),
                f"{data.get('revenue', 0):.2f}",
                data.get('records', 0),
                f"{avg_per_record:.2f}"
            ])
        
        writer.writerow([])
        
        # Payment Method Breakdown
        if 'payment_method_breakdown' in self.breakdown_data:
            writer.writerow(['PAYMENT METHOD BREAKDOWN'])
            writer.writerow(['Payment Method', 'Amount (₦)', 'Count'])
            
            payment_methods = self.breakdown_data['payment_method_breakdown']
            for method, data in payment_methods.items():
                writer.writerow([
                    method.replace('_', ' ').title(),
                    f"{data.get('amount', 0):.2f}",
                    data.get('count', 0)
                ])
        
        return output.getvalue()
    
    def export_excel(self):
        """
        Export revenue breakdown as Excel file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Revenue Summary"
        
        # Styling
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Header
        ws_summary['A1'] = "HMS Revenue Point Breakdown Analysis"
        ws_summary['A1'].font = header_font
        ws_summary['A2'] = f"Date Range: {self.date_range['start_date']} to {self.date_range['end_date']}"
        ws_summary['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary table
        row = 5
        ws_summary[f'A{row}'] = "Revenue Summary"
        ws_summary[f'A{row}'].font = subheader_font
        
        row += 1
        headers = ['Category', 'Revenue (₦)', 'Percentage (%)']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        
        row += 1
        summary = self.breakdown_data.get('summary_by_category', {})
        for category, data in summary.items():
            if category != 'grand_total':
                ws_summary.cell(row=row, column=1, value=category.replace('_', ' ').title())
                ws_summary.cell(row=row, column=2, value=float(data.get('revenue', 0)))
                ws_summary.cell(row=row, column=3, value=float(data.get('percentage', 0)))
                row += 1
        
        # Clinical Services Sheet
        ws_clinical = wb.create_sheet("Clinical Services")
        self._create_service_sheet(ws_clinical, 'Clinical Services', 
                                 self.breakdown_data.get('clinical_services', {}))
        
        # Support Services Sheet
        ws_support = wb.create_sheet("Support Services")
        self._create_service_sheet(ws_support, 'Support Services',
                                 self.breakdown_data.get('support_services', {}))
        
        # Specialty Departments Sheet
        ws_specialty = wb.create_sheet("Specialty Departments")
        self._create_specialty_sheet(ws_specialty, self.breakdown_data.get('specialty_departments', {}))
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_pdf(self):
        """
        Export revenue breakdown as PDF
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], 
                                   fontSize=18, spaceAfter=30, alignment=1)
        story.append(Paragraph("HMS Revenue Point Breakdown Analysis", title_style))
        
        # Date range and generation info
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, 
                                  spaceAfter=20, alignment=1)
        story.append(Paragraph(
            f"Date Range: {self.date_range['start_date']} to {self.date_range['end_date']}<br/>"
            f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", 
            info_style
        ))
        
        # Revenue Summary Table
        story.append(Paragraph("Revenue Summary", styles['Heading2']))
        
        summary_data = [['Category', 'Revenue (₦)', 'Percentage (%)']]
        summary = self.breakdown_data.get('summary_by_category', {})
        
        for category, data in summary.items():
            if category != 'grand_total':
                summary_data.append([
                    category.replace('_', ' ').title(),
                    f"₦{data.get('revenue', 0):,.2f}",
                    f"{data.get('percentage', 0):.1f}%"
                ])
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Clinical Services
        story.append(Paragraph("Clinical Services Breakdown", styles['Heading2']))
        
        clinical_data = [['Service', 'Revenue (₦)', 'Transactions', 'Avg Transaction (₦)']]
        clinical = self.breakdown_data.get('clinical_services', {})
        
        for service, data in clinical.items():
            clinical_data.append([
                service.title(),
                f"₦{data.get('revenue', 0):,.2f}",
                str(data.get('transactions', 0)),
                f"₦{data.get('avg_transaction', 0):,.2f}"
            ])
        
        clinical_table = Table(clinical_data)
        clinical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        
        story.append(clinical_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
    
    def _create_service_sheet(self, worksheet, title, services_data):
        """Create a service breakdown sheet in Excel"""
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14)
        
        headers = ['Service', 'Revenue (₦)', 'Transactions', 'Average Transaction (₦)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 4
        for service, data in services_data.items():
            worksheet.cell(row=row, column=1, value=service.title())
            worksheet.cell(row=row, column=2, value=float(data.get('revenue', 0)))
            worksheet.cell(row=row, column=3, value=data.get('transactions', 0))
            worksheet.cell(row=row, column=4, value=float(data.get('avg_transaction', 0)))
            row += 1
    
    def _create_specialty_sheet(self, worksheet, specialty_data):
        """Create specialty departments sheet in Excel"""
        worksheet['A1'] = "Specialty Departments"
        worksheet['A1'].font = Font(bold=True, size=14)
        
        headers = ['Department', 'Revenue (₦)', 'Records', 'Revenue per Record (₦)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 4
        for dept, data in specialty_data.items():
            avg_per_record = 0
            if data.get('records', 0) > 0:
                avg_per_record = data.get('revenue', 0) / data.get('records', 1)
            
            worksheet.cell(row=row, column=1, value=dept.replace('_', ' ').title())
            worksheet.cell(row=row, column=2, value=float(data.get('revenue', 0)))
            worksheet.cell(row=row, column=3, value=data.get('records', 0))
            worksheet.cell(row=row, column=4, value=float(avg_per_record))
            row += 1
    
    def _format_filters(self):
        """Format applied filters for display"""
        if not self.filters:
            return "None"
        
        filter_parts = []
        for key, value in self.filters.items():
            if value and value != 'all':
                filter_parts.append(f"{key.replace('_', ' ').title()}: {value}")
        
        return "; ".join(filter_parts) if filter_parts else "None"


class RevenueDataProcessor:
    """
    Processor for applying advanced filters to revenue data
    """
    
    def __init__(self, analyzer):
        self.analyzer = analyzer
    
    def apply_advanced_filters(self, filters):
        """
        Apply advanced filters to revenue data
        """
        # Get base breakdown data
        breakdown = self.analyzer.get_revenue_point_breakdown(include_trends=True)
        
        # Apply filters
        if filters.get('department_filter') and filters['department_filter'] != 'all':
            breakdown = self._filter_by_department(breakdown, filters['department_filter'])
        
        if filters.get('payment_method_filter') and filters['payment_method_filter'] != 'all':
            breakdown = self._filter_by_payment_method(breakdown, filters['payment_method_filter'])
        
        if filters.get('min_revenue') or filters.get('max_revenue'):
            breakdown = self._filter_by_revenue_range(
                breakdown, 
                filters.get('min_revenue'), 
                filters.get('max_revenue')
            )
        
        # Apply grouping
        if filters.get('group_by') and filters['group_by'] != 'none':
            breakdown = self._apply_grouping(breakdown, filters['group_by'])
        
        # Apply sorting
        if filters.get('sort_by'):
            breakdown = self._apply_sorting(breakdown, filters['sort_by'])
        
        return breakdown
    
    def _filter_by_department(self, breakdown, department):
        """Filter data by specific department"""
        # Implementation would depend on specific requirements
        return breakdown
    
    def _filter_by_payment_method(self, breakdown, payment_method):
        """Filter data by payment method"""
        # Implementation would depend on specific requirements
        return breakdown
    
    def _filter_by_revenue_range(self, breakdown, min_revenue, max_revenue):
        """Filter data by revenue range"""
        # Implementation would depend on specific requirements
        return breakdown
    
    def _apply_grouping(self, breakdown, group_by):
        """Apply grouping to data"""
        # Implementation would depend on specific requirements
        return breakdown
    
    def _apply_sorting(self, breakdown, sort_by):
        """Apply sorting to data"""
        # Implementation would depend on specific requirements
        return breakdown